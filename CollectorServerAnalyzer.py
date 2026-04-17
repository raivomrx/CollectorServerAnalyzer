import json
from openpyxl import Workbook

# -----------------------------
# FUNCTIONS
# -----------------------------

def load_data(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        with open(file_path, 'r', encoding='cp1252') as f:
            return json.load(f)


def analyze_security(data):
    risks = []
    score = 100

    # BitLocker
    if not data.get("Bitlocker-C"):
        risks.append(("BitLocker disabled", "HIGH"))
        score -= 20

    # Defender
    defender = data.get("Windows Defender", {})
    if defender.get("ProductState") != "On":
        risks.append(("Windows Defender not active", "HIGH"))
        score -= 20

    # Firewall
    firewall = data.get("Firewall") or []
    for profile in firewall:
        if profile.get("Enabled") != "True":
            risks.append((f"Firewall disabled ({profile.get('Profile')})", "HIGH"))
            score -= 10

    # Local admins
    admins = data.get("All_local_admins") or []
    if len(admins) > 2:
        risks.append(("Too many local administrators", "HIGH"))
        score -= 15

    # Risky services
    services = data.get("Non_standard_win_services") or []
    risky_tools = ["AnyDesk", "TeamViewer"]

    for svc in services:
        for tool in risky_tools:
            if tool.lower() in svc.get("Name", "").lower():
                risks.append((f"Remote access tool detected: {tool}", "HIGH"))
                score -= 15

    # Network
    for ip in data.get("IP_config" or []):
        ip_addr = ip.get("IPv4Address") or ""
        if ip_addr.startswith("169.254"):
            risks.append(("APIPA address detected (network issue)", "MEDIUM"))
            score -= 5

    return risks, max(score, 0)


def export_excel(data, risks, score, filename="report.xlsx"):
    wb = Workbook()

    # Overview
    ws1 = wb.active
    ws1.title = "Overview"
    ws1.append(["ComputerName", data.get("ComputerName")])
    ws1.append(["OS", data.get("OS")])
    ws1.append(["Security Score", score])

    # Risks
    ws2 = wb.create_sheet(title="Risks")
    ws2.append(["Issue", "Severity"])
    for r in risks:
        ws2.append(r)

    # Software
    ws3 = wb.create_sheet(title="Software")
    ws3.append(["Name", "Version"])
    for sw in data.get("Software" or []):
        if sw.get("Name") and sw.get("Version"):
            ws3.append([sw.get("Name"), sw.get("Version")])

    wb.save(filename)


# -----------------------------
# MAIN
# -----------------------------

import os

if __name__ == "__main__":
    for file in os.listdir():
        if file.endswith(".json"):
            print(f"Töötlen faili: {file}")

            data = load_data(file)
            risks, score = analyze_security(data)

            output_file = file.replace(".json", ".xlsx")
            export_excel(data, risks, score, output_file)

    print("Kõik failid töödeldud!")